# -*- coding: utf-8 -*-
"""
전표 5400004827 SAP 다운로드 → 필터링 없이 Excel 저장
VBScript 1:1 변환 (추가 다이얼로그 처리 없음)
"""
import sys, os, time, glob, shutil
sys.stdout.reconfigure(encoding='utf-8')
import win32com.client

# ── SAP 연결 ──────────────────────────────────────────
sap_gui    = win32com.client.GetObject("SAPGUI")
application = sap_gui.GetScriptingEngine
connection  = application.Children(0)
session     = connection.Children(0)
print(f"SAP: {session.Info.SystemName} / {session.Info.User}")

# 저장 파일 감지용 스냅샷
desktop     = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")
userprofile = os.environ.get("USERPROFILE", "")
watch_dirs  = [
    desktop,
    os.environ.get("TEMP", ""),
    os.path.join(userprofile, "Documents"),
    os.path.join(userprofile, "AppData", "Local", "Temp"),
    os.getcwd(),
]
before = {}
for d in watch_dirs:
    if d and os.path.isdir(d):
        for pat in ("*.xls", "*.XLS", "*.xlsx"):
            for f in glob.glob(os.path.join(d, pat)):
                try: before[f] = os.path.getmtime(f)
                except: pass
t0 = time.time()

# ── VBScript 1:1 변환 ─────────────────────────────────
session.findById("wnd[0]").maximize()
session.findById("wnd[0]/usr/ctxtKBKP-BELNR").text = ""
session.findById("wnd[0]").sendVKey(4)
time.sleep(1.5)

TAB = ("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001"
       "/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220")

fld5 = session.findById(TAB + "/txtG_SELFLD_TAB-LOW[5,24]")
fld5.text = ""; fld5.setFocus(); fld5.caretPosition = 0
session.findById("wnd[1]").sendVKey(0)
time.sleep(1.0)

fld4 = session.findById(TAB + "/txtG_SELFLD_TAB-LOW[4,24]")
fld4.text = ""; fld4.setFocus(); fld4.caretPosition = 0
session.findById("wnd[1]/tbar[0]/btn[0]").press()
time.sleep(2.0)

session.findById("wnd[1]/usr/lbl[1,12]").setFocus()
session.findById("wnd[1]/usr/lbl[1,12]").caretPosition = 8
session.findById("wnd[1]").sendVKey(2)
time.sleep(1.0)

session.findById("wnd[0]").sendVKey(0)
time.sleep(3.0)
print("상세 화면 로딩 완료")

session.findById("wnd[0]/mbar/menu[4]/menu[5]/menu[2]/menu[2]").select()
time.sleep(1.0)

RADIO = ("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150"
         "/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]")
session.findById(RADIO).select()
session.findById(RADIO).setFocus()
session.findById("wnd[1]/tbar[0]/btn[0]").press()   # OK — VBScript 끝
time.sleep(3.0)

# 저장 다이얼로그가 뜰 경우 Enter 로 승인 (sap_download.py 방식)
shell = win32com.client.Dispatch("WScript.Shell")
shell.SendKeys("{ENTER}", 0)
time.sleep(3.0)

# ── 저장된 파일 감지 ──────────────────────────────────
found = None
for d in watch_dirs:
    if not d or not os.path.isdir(d): continue
    for pat in ("*.xls", "*.XLS", "*.xlsx"):
        for f in glob.glob(os.path.join(d, pat)):
            try:
                mt = os.path.getmtime(f)
                if mt >= t0 or (f in before and mt > before[f]) or f not in before:
                    found = f
            except: pass

if not found:
    print("[경고] 파일 감지 실패 — 바탕화면을 확인하세요")
    sys.exit(1)

print(f"다운로드: {found}  ({os.path.getsize(found):,} bytes)")

# ── RAW 데이터 그대로 Excel 저장 ──────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src", "modules", "analytics"))
from mr11_processor import _read_sap_xls
import openpyxl
from openpyxl.utils import get_column_letter

rows = _read_sap_xls(found)
print(f"읽은 행수: {len(rows)}")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "5400004827"

for r_idx, row in enumerate(rows, 1):
    for c_idx, val in enumerate(row, 1):
        ws.cell(r_idx, c_idx).value = val

for col in ws.columns:
    max_len = max((len(str(cell.value or "")) for cell in col), default=0)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

out = os.path.join(os.path.dirname(__file__), "5400004827_raw.xlsx")
wb.save(out)
print(f"저장 완료: {out}")
print(f"  {ws.max_row}행 x {ws.max_column}열")
