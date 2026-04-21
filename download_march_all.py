# -*- coding: utf-8 -*-
"""
MR11SHOW 월별 전표 전체 다운로드 → 1개 Excel 저장

사용법:
    python download_march_all.py --month 4            # 2026년 4월
    python download_march_all.py --month 4 --year 2026
    python download_march_all.py -m 4 -y 2026
"""
import sys, re, time, argparse
sys.stdout.reconfigure(encoding='utf-8')
import win32com.client, win32clipboard
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(description="MR11SHOW 월별 전표 일괄 다운로드")
parser.add_argument("--year",  "-y", default="2026", help="회계연도 (기본값: 2026)")
parser.add_argument("--month", "-m", required=True,  help="대상 월 숫자 (예: 3, 4 ...)")
args = parser.parse_args()

FISCAL_YEAR  = args.year
TARGET_MONTH = args.month.zfill(2)   # '4' → '04'

# ── SAP 연결 ──────────────────────────────────────────
sap_gui = win32com.client.GetObject("SAPGUI")
session = sap_gui.GetScriptingEngine.Children(0).Children(0)
print(f"SAP: {session.Info.SystemName} / {session.Info.User}")

session.findById("wnd[0]/tbar[0]/okcd").text = "/nMR11SHOW"
session.findById("wnd[0]").sendVKey(0)
time.sleep(1.5)

# MR11SHOW 첫 화면: 회계연도 입력
session.findById("wnd[0]/usr/txtKBKP-GJAHR").text = FISCAL_YEAR
print(f"  회계연도: {FISCAL_YEAR}")

# ── Phase 1: F4 matchcode → 전표 목록 읽기 ───────────
print("\n[1/3] 전표 목록 조회 중...")
TAB = ("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001"
       "/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220")

session.findById("wnd[0]").maximize()
session.findById("wnd[0]/usr/ctxtKBKP-BELNR").text = ""
session.findById("wnd[0]").sendVKey(4)
time.sleep(1.5)
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[5,24]").text = ""
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[5,24]").setFocus()
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[5,24]").caretPosition = 0
session.findById("wnd[1]").sendVKey(0)
time.sleep(1.0)
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[4,24]").text = ""
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[4,24]").setFocus()
session.findById(TAB + "/txtG_SELFLD_TAB-LOW[4,24]").caretPosition = 0
session.findById("wnd[1]/tbar[0]/btn[0]").press()
time.sleep(2.0)

# 팝업에서 모든 레이블 읽기 (여러 페이지 대응)
def read_popup_labels():
    usr = session.findById("wnd[1]/usr")
    labels = {}
    for i in range(usr.Children.Count):
        lbl = usr.Children.ElementAt(i)
        try:
            lid = lbl.Id
            bracket = lid[lid.rfind("[") + 1: lid.rfind("]")]
            c, r = map(int, bracket.split(","))
            labels[(c, r)] = lbl.Text
        except:
            pass
    return labels

all_docs = {}   # {doc_no: posting_date}
prev_first = None

for page in range(50):
    labels = read_popup_labels()
    if not labels:
        break

    # 첫 행으로 페이지 중복 감지
    first_val = labels.get((1, 1), "").strip()
    if first_val == prev_first and page > 0:
        break
    prev_first = first_val

    # 행별 데이터 수집
    rows = {}
    for (c, r), text in labels.items():
        t = text.strip()
        if t:
            rows.setdefault(r, {})[c] = t

    for r, cols in rows.items():
        # col 1 = 전표번호 (7자리 이상 숫자)
        doc = cols.get(1, "").replace(" ", "")
        if not (len(doc) >= 7 and doc.isdigit()):
            continue
        # 날짜: YYYY.MM.DD 패턴이 있는 컬럼
        all_dates = []
        for k in sorted(cols.keys()):
            m = re.search(r"\d{4}\.\d{2}\.\d{2}", cols[k])
            if m:
                all_dates.append(m.group())
        # 회계연도와 일치하는 날짜 우선 (연말전표: 전기일=2025.12 / 입력일=2026.01)
        date = next((d for d in all_dates if d.startswith(FISCAL_YEAR)), "")
        if not date and all_dates:
            date = all_dates[0]
        if doc not in all_docs:
            all_docs[doc] = date

    # 다음 페이지
    try:
        session.findById("wnd[1]").sendVKey(77)  # Page Down
        time.sleep(0.5)
    except:
        break

# 팝업 닫기
session.findById("wnd[1]").sendVKey(12)
time.sleep(0.5)

print(f"  전체 전표: {len(all_docs)}개")

# 3월 전표 필터링
march_docs = {doc: date for doc, date in all_docs.items()
              if f"{FISCAL_YEAR}.{TARGET_MONTH}" in date}

# 날짜 없는 전표도 포함 여부 확인
undated = {doc: date for doc, date in all_docs.items() if not date}
if undated:
    print(f"  날짜 미확인: {len(undated)}개 → 제외")

print(f"  3월 전표: {len(march_docs)}개")
for doc, date in sorted(march_docs.items()):
    print(f"    {doc}  ({date})")

if not march_docs:
    print("3월 전표 없음")
    sys.exit(0)


# ── Phase 2: 전표별 상세 내보내기 (클립보드) ──────────
print(f"\n[2/3] 전표별 상세 추출 중...")
RADIO = ("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150"
         "/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]")


def export_clipboard(doc_no):
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nMR11SHOW"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1.0)
    session.findById("wnd[0]/usr/txtKBKP-GJAHR").text = FISCAL_YEAR
    session.findById("wnd[0]/usr/ctxtKBKP-BELNR").text = doc_no
    session.findById("wnd[0]").sendVKey(0)   # Enter → 상세 화면
    time.sleep(2.5)
    session.findById("wnd[0]/mbar/menu[4]/menu[5]/menu[2]/menu[2]").select()
    time.sleep(1.0)
    session.findById(RADIO).select()
    session.findById(RADIO).setFocus()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(2.0)
    win32clipboard.OpenClipboard()
    try:
        data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
    finally:
        win32clipboard.CloseClipboard()
    return data


# ── Phase 3: 클립보드 파싱 ────────────────────────────
def clean_num(s):
    s = str(s).strip().replace(",", "")
    if s.endswith("-"):
        try: return -float(s[:-1])
        except: return s
    try: return float(s)
    except: return s


def parse_a_row(content):
    m = re.match(r"^(\d{10})\s+(\d+)\s+(\d{4}\.\d{2}\.\d{2})\s+(.*)", content.strip())
    if not m:
        return None
    po, item, date, rest = m.group(1), m.group(2), m.group(3), m.group(4).strip()
    parts = [p.strip() for p in re.split(r"\s{2,}", rest) if p.strip()]
    name, plnt, desc, oun = "", "", "", ""
    if len(parts) >= 4:
        name = parts[0]
        plnt_desc = parts[2]
        oun = parts[-1]
        pd = plnt_desc.split(None, 1)
        plnt = pd[0] if pd else ""
        desc = pd[1] if len(pd) > 1 else ""
    elif len(parts) == 3:
        name = parts[0]
        pd = parts[1].split(None, 1)
        plnt = pd[0] if pd else ""
        desc = pd[1] if len(pd) > 1 else ""
        oun = parts[2]
    elif parts:
        name = parts[0]
    return {"po": po, "item": item, "date": date,
            "name": name, "plnt": plnt, "desc": desc, "oun": oun}


def parse_clipboard_data(data, doc_no):
    lines = data.splitlines()

    def is_data(line):
        if not line.startswith("|"): return False
        c = line[1:].strip()
        if not c or c.startswith("-"): return False
        if re.match(r"^[가-힣]", c): return False
        return True

    data_lines = [l[1:].rstrip().rstrip("|").rstrip() for l in lines if is_data(l)]
    records = []
    i = 0
    while i < len(data_lines):
        a = data_lines[i].strip()
        if re.match(r"^4\d{9}\s", a):
            rec_a = parse_a_row(a)
            i += 1
            if rec_a:
                while i < len(data_lines):
                    b = data_lines[i].strip()
                    if re.match(r"^4\d{9}\s", b): break
                    b_m = re.match(r"^\d+\s+(\d{10})\s+(\d+)\s+(.*)", b)
                    if b_m:
                        brest = b_m.group(3).strip()
                        parts = re.split(r"\s{2,}", brest)
                        acct = parts[0].strip()
                        nums = [p.strip() for p in parts[1:]
                                if re.match(r"^[\d,]+-?$", p.strip())]
                        records.append({
                            "전표번호":   doc_no,
                            "구매 문서":  rec_a["po"],
                            "품목":       rec_a["item"],
                            "PO 일자":    rec_a["date"],
                            "계정키이름": acct,
                            "이름 1":     rec_a["name"],
                            "차이 수량":  clean_num(nums[0]) if nums else "",
                            "차이 금액":  clean_num(nums[1]) if len(nums) > 1 else "",
                            "Plnt":       rec_a["plnt"],
                            "내역":       rec_a["desc"],
                            "OUn":        rec_a["oun"],
                        })
                    i += 1
        else:
            i += 1
    return records


all_records = []
for idx, (doc_no, post_date) in enumerate(sorted(march_docs.items()), 1):
    print(f"  [{idx}/{len(march_docs)}] 전표 {doc_no} ({post_date}) ...", end=" ", flush=True)
    try:
        clip_data = export_clipboard(doc_no)
        recs = parse_clipboard_data(clip_data, doc_no)
        all_records.extend(recs)
        print(f"{len(recs)}건")
    except Exception as e:
        print(f"오류: {e}")

print(f"\n  총 {len(all_records)}건 추출")


# ── Phase 4: Excel 저장 ───────────────────────────────
print("\n[3/3] Excel 저장 중...")
COLS = ["전표번호", "구매 문서", "품목", "PO 일자", "계정키이름", "이름 1",
        "차이 수량", "차이 금액", "Plnt", "내역", "OUn"]
BORDER = Border(
    left=Side(style="thin", color="AAAAAA"), right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),  bottom=Side(style="thin", color="AAAAAA"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"MR11_{TARGET_MONTH}월"

for c in range(1, len(COLS) + 1):
    ws.cell(1, c).value = ""
last_row = len(all_records) + 2
ws.cell(1, 8).value = f"=SUM(H3:H{last_row})"
ACCT_FMT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
ws.cell(1, 8).number_format = ACCT_FMT
ws.cell(1, 8).alignment = Alignment(horizontal="right", vertical="center")

for ci, col in enumerate(COLS, 1):
    cell = ws.cell(2, ci, value=col)
    cell.fill = PatternFill("solid", fgColor="BFBFBF")
    cell.font = Font(bold=True, size=10)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

for ri, rec in enumerate(all_records, 3):
    fill = PatternFill("solid", fgColor="F2F2F2" if ri % 2 == 0 else "FFFFFF")
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(ri, ci, value=rec.get(col, ""))
        cell.fill = fill
        cell.font = Font(size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ci in (7, 8):  # G=차이 수량, H=차이 금액
            cell.number_format = ACCT_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[ri].height = 16

for col in ws.columns:
    max_len = max((len(str(cell.value or "")) for cell in col), default=0)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
ws.freeze_panes = "A3"

import os
from datetime import datetime
out = os.path.join(os.path.dirname(__file__),
                   f"MR11_{FISCAL_YEAR}_{TARGET_MONTH}월반제리스트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
wb.save(out)
print(f"\n{'='*55}")
print(f"  저장 완료: {os.path.basename(out)}")
print(f"  전표 수  : {len(march_docs)}개")
print(f"  데이터   : {len(all_records)}건")
print(f"  차이금액 : {total:,.0f}")
print(f"{'='*55}")
