# -*- coding: utf-8 -*-
"""
MR11 반제리스트 다운로더 - Windows GUI
"""
import sys, re, time, os, threading
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog

def _icon_path():
    base = sys._MEIPASS if hasattr(sys, "_MEIPASS") else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "duck.ico")

def _default_save_dir():
    if hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ── 색상 팔레트 ───────────────────────────────────────────
BG        = "#F0F4F8"   # 전체 배경 (연한 청회색)
CARD      = "#FFFFFF"   # 카드 배경
HEADER_BG = "#1E3A5F"   # 헤더 배경 (딥 네이비)
HEADER_FG = "#FFFFFF"
ACCENT    = "#2563EB"   # 버튼·포인트 (블루)
ACCENT_HV = "#1D4ED8"   # 호버
BTN_FG    = "#FFFFFF"
LABEL_FG  = "#374151"   # 일반 레이블
MUTED     = "#6B7280"   # 보조 텍스트
BORDER    = "#D1D5DB"   # 테두리
LOG_BG    = "#0F172A"   # 로그 배경 (다크 네이비)
LOG_FG    = "#FFFFFF"   # 로그 기본 텍스트

# ── GUI 앱 ────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MR11 반제리스트 다운로더")
        self.resizable(True, True)
        self.minsize(560, 580)
        self.configure(bg=BG)
        try:
            self.iconbitmap(_icon_path())
        except Exception:
            pass
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 560, 580
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _card(self, parent, **kw):
        f = tk.Frame(parent, bg=CARD, relief="flat",
                     highlightbackground=BORDER, highlightthickness=1, **kw)
        return f

    def _label(self, parent, text, bold=False, size=10, fg=LABEL_FG):
        font = ("맑은 고딕", size, "bold" if bold else "normal")
        return tk.Label(parent, text=text, font=font, fg=fg, bg=parent["bg"])

    def _build_ui(self):
        # ── 헤더 ──
        hdr = tk.Frame(self, bg=HEADER_BG, height=64)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="MR11  반제리스트 다운로더",
                 font=("맑은 고딕", 15, "bold"),
                 fg=HEADER_FG, bg=HEADER_BG).place(relx=0.5, rely=0.5, anchor="center")

        # ── 조회 조건 카드 ──
        card1 = self._card(self)
        card1.pack(fill="x", padx=20, pady=(16, 0))

        self._label(card1, "조회 조건", bold=True, size=10, fg=MUTED).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=14, pady=(10, 4))

        self._label(card1, "회계연도").grid(row=1, column=0, padx=(14,6), pady=8, sticky="e")
        self.year_var = tk.StringVar(value="2026")
        tk.Entry(card1, textvariable=self.year_var, width=9,
                 font=("맑은 고딕", 11), relief="solid",
                 highlightbackground=BORDER, highlightthickness=1
                 ).grid(row=1, column=1, sticky="w", pady=8)

        self._label(card1, "전기월").grid(row=1, column=2, padx=(20,6), pady=8, sticky="e")
        self.month_var = tk.StringVar(value="3")
        style = ttk.Style()
        style.configure("Custom.TCombobox", padding=4)
        ttk.Combobox(card1, textvariable=self.month_var,
                     values=[str(i) for i in range(1, 13)],
                     width=6, state="readonly",
                     style="Custom.TCombobox"
                     ).grid(row=1, column=3, sticky="w", pady=8, padx=(0,14))

        # ── 저장 위치 카드 ──
        card2 = self._card(self)
        card2.pack(fill="x", padx=20, pady=(10, 0))

        self._label(card2, "저장 위치", bold=True, size=10, fg=MUTED).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=14, pady=(10, 4))

        self.save_dir_var = tk.StringVar(value=_default_save_dir())
        dir_entry = tk.Entry(card2, textvariable=self.save_dir_var,
                             font=("맑은 고딕", 9), relief="solid",
                             highlightbackground=BORDER, highlightthickness=1,
                             state="readonly", readonlybackground="#F9FAFB",
                             fg=LABEL_FG, width=44)
        dir_entry.grid(row=1, column=0, padx=(14, 6), pady=(0, 12), sticky="w")

        tk.Button(card2, text="폴더 선택", font=("맑은 고딕", 9),
                  bg="#E5E7EB", fg=LABEL_FG, relief="flat", cursor="hand2",
                  activebackground=BORDER, padx=8, pady=4,
                  command=self._pick_dir).grid(row=1, column=1, pady=(0, 12), padx=(0,14))

        # ── 실행 버튼 ──
        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(pady=14)
        self.run_btn = tk.Button(btn_frame, text="▶   실행",
                                 font=("맑은 고딕", 12, "bold"),
                                 bg=ACCENT, fg=BTN_FG,
                                 activebackground=ACCENT_HV, activeforeground=BTN_FG,
                                 disabledforeground=BTN_FG,
                                 relief="flat", cursor="hand2",
                                 padx=36, pady=10,
                                 command=self._on_run)
        self.run_btn.pack()

        # ── 로그 카드 ──
        card3 = self._card(self)
        card3.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        self._label(card3, "진행 상황", bold=True, size=10, fg=MUTED).pack(
            anchor="w", padx=14, pady=(10, 4))

        self.log_box = scrolledtext.ScrolledText(
            card3, height=12, state="disabled",
            font=("Consolas", 9), bg=LOG_BG, fg=LOG_FG,
            insertbackground="white", relief="flat",
            borderwidth=0, padx=8, pady=6)
        self.log_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _pick_dir(self):
        d = filedialog.askdirectory(title="저장 폴더 선택",
                                    initialdir=self.save_dir_var.get())
        if d:
            self.save_dir_var.set(d)

    def _log(self, msg):
        def _write():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _write)

    def _on_run(self):
        year  = self.year_var.get().strip()
        month = self.month_var.get().strip()
        if not year.isdigit() or not month.isdigit():
            messagebox.showerror("입력 오류", "회계연도와 전기월을 올바르게 입력하세요.")
            return

        messagebox.showinfo(
            "SAP 로그인 확인",
            "SAP GUI에 먼저 로그인되어 있어야 합니다.\n\n"
            f"  · 회계연도 : {year}년\n"
            f"  · 전기월   : {month}월\n\n"
            "로그인 확인 후 확인을 눌러주세요.",
        )

        self.run_btn.config(state="disabled", text="실행 중...")
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

        save_dir = self.save_dir_var.get()
        thread = threading.Thread(target=self._run_task,
                                  args=(year, month, save_dir), daemon=True)
        thread.start()

    def _run_task(self, fiscal_year, month_str, save_dir):
        import pythoncom
        pythoncom.CoInitialize()
        try:
            run_download(fiscal_year, month_str, save_dir, self._log)
        except Exception as e:
            self._log(f"\n[오류] {e}")
        finally:
            pythoncom.CoUninitialize()
            self.after(0, lambda: self.run_btn.config(state="normal", text="▶   실행"))


# ── 다운로드 로직 ─────────────────────────────────────────
def run_download(FISCAL_YEAR, month_str, save_dir, log):
    import win32com.client, win32clipboard
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    TARGET_MONTH = month_str.zfill(2)
    ACCT_FMT = '_(* #,##0_);_(* -#,##0_);_(* "-"_);_(@_)'
    BORDER = Border(
        left=Side(style="thin", color="AAAAAA"), right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),  bottom=Side(style="thin", color="AAAAAA"),
    )
    COLS = ["전표번호", "구매 문서", "품목", "PO 일자", "계정키이름", "이름 1",
            "차이 수량", "차이 금액", "Plnt", "내역", "OUn"]

    # ── SAP 연결 ──
    shell_obj = win32com.client.Dispatch("WScript.Shell")
    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        log(f"SAP 연결: {session.Info.SystemName} / {session.Info.User}")
    except Exception as e:
        log(f"[오류] SAP 연결 실패: {e}")
        log("SAP GUI가 실행 중이고 로그인되어 있는지 확인하세요.")
        return

    # ── MR11SHOW 이동 + 회계연도 입력 ──
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nMR11SHOW"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1.5)
    session.findById("wnd[0]/usr/txtKBKP-GJAHR").text = FISCAL_YEAR
    log(f"회계연도 {FISCAL_YEAR} 입력 완료")

    # ── Phase 1: F4 matchcode → 전표 목록 ──
    log(f"\n[1/3] {FISCAL_YEAR}년 {TARGET_MONTH}월 전표 목록 조회 중...")
    TAB_BASE = ("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001"
                "/ssubSUBSCR_PRESEL:SAPLSDH4:0220")
    TAB = TAB_BASE + "/sub:SAPLSDH4:0220"

    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/usr/ctxtKBKP-BELNR").text = ""
    session.findById("wnd[0]").sendVKey(4)
    time.sleep(1.5)

    # MAXRECORDS 제한 해제 (9999)
    session.findById(TAB_BASE + "/txtDDSHF4CTRL-MAXRECORDS").text = "9999"
    # 회계연도 필터 + 날짜 필터 초기화
    session.findById(TAB + "/txtG_SELFLD_TAB-LOW[1,24]").text = FISCAL_YEAR
    session.findById(TAB + "/txtG_SELFLD_TAB-LOW[4,24]").text = ""
    session.findById(TAB + "/txtG_SELFLD_TAB-LOW[5,24]").text = ""
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(2.0)

    # matchcode_popup_debug.txt 분석:
    # 팝업 레이블 = lbl[col, row] 형식
    # col 1=전표번호, col 12=연도, col 17=전기일, col 28=입력일
    # 헤더=row 1, row 2 없음(gap), 데이터=row 3~
    # → usr.Children 미사용(SAP 스크롤 포커스 유지), findById 직접 접근
    def read_page_direct():
        result = {}
        consecutive_fails = 0
        for row in range(1, 100):
            try:
                c1 = session.findById(f"wnd[1]/usr/lbl[1,{row}]").Text.strip()
                consecutive_fails = 0
                if not c1:
                    continue
                row_data = {1: c1}
                for col in (12, 17, 28):
                    try:
                        t = session.findById(f"wnd[1]/usr/lbl[{col},{row}]").Text.strip()
                        if t:
                            row_data[col] = t
                    except Exception:
                        pass
                result[row] = row_data
            except Exception:
                consecutive_fails += 1
                if consecutive_fails >= 5:
                    break
        return result

    all_docs = {}
    prev_first = None
    for page in range(100):
        page_data = read_page_direct()
        data_rows = {r: d for r, d in page_data.items() if r >= 3}
        if not data_rows:
            break
        first_doc = data_rows.get(min(data_rows.keys()), {}).get(1, "").strip()
        if first_doc == prev_first and page > 0:
            break
        prev_first = first_doc
        for r, cols in data_rows.items():
            doc = cols.get(1, "").replace(" ", "")
            if not (len(doc) >= 7 and doc.isdigit()):
                continue
            # 연말 전표 대응: 전기일=2025.12.31 / 입력일=2026.01.xx 혼재
            # → 회계연도와 일치하는 날짜 우선 선택
            all_dates = [d for d in [cols.get(17, ""), cols.get(28, "")] if d]
            date = next((d for d in all_dates if d.startswith(FISCAL_YEAR)), "")
            if not date and all_dates:
                date = all_dates[0]
            if doc not in all_docs:
                all_docs[doc] = date
        try:
            session.findById("wnd[1]").sendVKey(82)
            time.sleep(0.5)
        except Exception:
            break

    session.findById("wnd[1]").sendVKey(12)
    time.sleep(0.5)

    target_docs = {doc: date for doc, date in all_docs.items()
                   if f"{FISCAL_YEAR}.{TARGET_MONTH}" in date}
    log(f"  전체 전표: {len(all_docs)}개 / {TARGET_MONTH}월: {len(target_docs)}개")
    for doc, date in sorted(target_docs.items()):
        log(f"    {doc}  ({date})")

    if not target_docs:
        log(f"{TARGET_MONTH}월 전표 없음 — 종료")
        return

    # ── Phase 2: 전표별 클립보드 내보내기 ──
    log(f"\n[2/3] 전표별 상세 추출 중...")
    RADIO = ("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150"
             "/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]")

    def export_clipboard(doc_no):
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nMR11SHOW"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1.0)
        session.findById("wnd[0]/usr/txtKBKP-GJAHR").text = FISCAL_YEAR
        session.findById("wnd[0]/usr/ctxtKBKP-BELNR").text = doc_no
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(2.5)
        session.findById("wnd[0]/mbar/menu[4]/menu[5]/menu[2]/menu[2]").select()
        time.sleep(1.0)
        session.findById(RADIO).select()
        session.findById(RADIO).setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(2.0)
        win32clipboard.OpenClipboard()
        try:
            data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        finally:
            win32clipboard.CloseClipboard()
        return data

    def clean_num(s):
        s = str(s).strip().replace(",", "")
        if s.endswith("-"):
            try: return -float(s[:-1])
            except: return s
        try: return float(s)
        except: return s

    def parse_a_row(content):
        m = re.match(r"^(\d{10})\s+(\d+)\s+(\d{4}\.\d{2}\.\d{2})\s+(.*)", content.strip())
        if not m:
            return None
        po, item, date, rest = m.group(1), m.group(2), m.group(3), m.group(4).strip()
        parts = [p.strip() for p in re.split(r"\s{2,}", rest) if p.strip()]
        name, plnt, desc, oun = "", "", "", ""
        if not parts:
            pass
        elif len(parts) == 1:
            name = parts[0]
        else:
            name = parts[0]
            last_tok = parts[-1].split()
            if last_tok and re.match(r'^\d{4}$', last_tok[0]):
                # Case C: parts[-1] 자체가 '1000 내역 OUn' 형태로 묶인 경우
                plnt = last_tok[0]
                oun  = last_tok[-1] if len(last_tok) > 1 else ""
                desc = " ".join(last_tok[1:-1]) if len(last_tok) > 2 else ""
            else:
                oun = parts[-1]
                # Plnt: 4자리 숫자 코드(1000, 2000 등)로 시작하는 part 탐색
                # Case A: '1000 내역설명' → 같은 part에 Plnt+내역
                # Case B: '1000', '내역설명' → 별도 part로 분리된 경우
                for idx in range(1, len(parts) - 1):
                    tok = parts[idx].split(None, 1)
                    if tok and re.match(r'^\d{4}$', tok[0]):
                        plnt = tok[0]
                        if len(tok) > 1:
                            desc = tok[1]           # Case A: 같은 part에 내역 있음
                        elif idx + 1 <= len(parts) - 2:
                            desc = parts[idx + 1]   # Case B: 바로 다음 part가 내역
                        break
        return {"po": po, "item": item, "date": date,
                "name": name, "plnt": plnt, "desc": desc, "oun": oun}

    def parse_clipboard_data(data, doc_no):
        lines = data.splitlines()
        def is_data(line):
            if not line.startswith("|"): return False
            c = line[1:].strip()
            if not c or c.startswith("-"): return False
            if re.match(r"^[가-힣]", c): return False
            return True
        data_lines = [l[1:].rstrip().rstrip("|").rstrip() for l in lines if is_data(l)]
        records = []
        i = 0
        while i < len(data_lines):
            a = data_lines[i].strip()
            if re.match(r"^4\d{9}\s", a):
                rec_a = parse_a_row(a)
                i += 1
                if rec_a:
                    while i < len(data_lines):
                        b = data_lines[i].strip()
                        if re.match(r"^4\d{9}\s", b): break
                        b_m = re.match(r"^\d+\s+(\d{10})\s+(\d+)\s+(.*)", b)
                        if b_m:
                            brest = b_m.group(3).strip()
                            parts = re.split(r"\s{2,}", brest)
                            acct = parts[0].strip()
                            nums = [p.strip() for p in parts[1:]
                                    if re.match(r"^[\d,]+-?$", p.strip())]
                            records.append({
                                "전표번호":   doc_no,
                                "구매 문서":  rec_a["po"],
                                "품목":       rec_a["item"],
                                "PO 일자":    rec_a["date"],
                                "계정키이름": acct,
                                "이름 1":     rec_a["name"],
                                "차이 수량":  clean_num(nums[0]) if nums else "",
                                "차이 금액":  clean_num(nums[1]) if len(nums) > 1 else "",
                                "Plnt":       rec_a["plnt"],
                                "내역":       rec_a["desc"],
                                "OUn":        rec_a["oun"],
                            })
                        i += 1
            else:
                i += 1
        return records

    all_records = []
    for idx, (doc_no, post_date) in enumerate(sorted(target_docs.items()), 1):
        log(f"  [{idx}/{len(target_docs)}] 전표 {doc_no} ({post_date}) ...", )
        try:
            clip_data = export_clipboard(doc_no)
            recs = parse_clipboard_data(clip_data, doc_no)
            all_records.extend(recs)
            log(f"    → {len(recs)}건 추출")
        except Exception as e:
            log(f"    → 오류: {e}")

    log(f"\n  총 {len(all_records)}건 추출")

    # ── Phase 3: Excel 저장 ──
    log("\n[3/3] Excel 저장 중...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"MR11_{TARGET_MONTH}월"

    for c in range(1, len(COLS) + 1):
        ws.cell(1, c).value = ""
    last_row = len(all_records) + 2
    ws.cell(1, 8).value = f"=SUM(H3:H{last_row})"
    ws.cell(1, 8).number_format = ACCT_FMT
    ws.cell(1, 8).alignment = Alignment(horizontal="right", vertical="center")

    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(2, ci, value=col)
        cell.fill = PatternFill("solid", fgColor="BFBFBF")
        cell.font = Font(bold=True, size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    for ri, rec in enumerate(all_records, 3):
        fill = PatternFill("solid", fgColor="F2F2F2" if ri % 2 == 0 else "FFFFFF")
        for ci, col in enumerate(COLS, 1):
            cell = ws.cell(ri, ci, value=rec.get(col, ""))
            cell.fill = fill
            cell.font = Font(size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (7, 8):
                cell.number_format = ACCT_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[ri].height = 16

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws.column_dimensions["K"].width = 15
    ws.freeze_panes = "A3"

    out = os.path.join(save_dir,
                       f"MR11반제리스트_{FISCAL_YEAR}_{TARGET_MONTH}월.xlsx")
    wb.save(out)

    log(f"\n{'='*45}")
    log(f"  저장 완료: {os.path.basename(out)}")
    log(f"  전표 수  : {len(target_docs)}개")
    log(f"  데이터   : {len(all_records)}건")
    log(f"{'='*45}")

    messagebox.showinfo("완료", f"완료!\n\n저장 파일: {os.path.basename(out)}\n전표 수: {len(target_docs)}개 / 데이터: {len(all_records)}건")


# ── 진입점 ────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
