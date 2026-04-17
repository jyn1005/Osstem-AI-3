"""
MR11SHOW rawdata 처리 스크립트
SAP MR11SHOW 상세 화면에서 다운받은 CSV를 정리하여
반제리스트 Excel 마스터 파일에 누적 추가합니다.

사용법:
    # 파일 1개 처리
    python src/modules/analytics/mr11_processor.py --input rawdata.csv --master 반제리스트.xlsx

    # 폴더 내 CSV 전체 처리
    python src/modules/analytics/mr11_processor.py --input ./rawdata폴더/ --master 반제리스트.xlsx
"""

import sys
import os
import glob
import argparse
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────
HEADER_COLOR  = "BFBFBF"   # 헤더 배경 (회색)
HEADER_FONT   = "000000"   # 헤더 글자 (검정)
ALT_ROW_COLOR = "F2F2F2"   # 짝수 행 배경 (연한 회색)
NEW_ROW_COLOR = "E2EFDA"   # 신규 추가 행 배경 (연한 초록) — 추가분 구별용

# 마스터 파일 컬럼 순서
MASTER_COLS = [
    "전표번호", "구매 문서", "품목", "PO 일자",
    "계정키이름", "이름 1", "차이 수량", "차이 금액",
    "Plnt", "내역", "OUn",
]
# 합계금액이 위치하는 컬럼 인덱스 (1-based, Excel H열 = 8번째)
TOTAL_COL_IDX = 8   # "차이 금액" 컬럼
# ──────────────────────────────────────────────────────


def _clean_val(v) -> str:
    """단일 셀 값을 클린 문자열로 변환합니다."""
    s = str(v).strip()
    return "" if s in ("nan", "None", "") else s


def _clean_row(row_values) -> list:
    """pandas row를 클린 문자열 리스트로 변환합니다."""
    return [_clean_val(v) for v in row_values]


def _to_number(text: str) -> float:
    """숫자 문자열을 float으로 변환합니다. 쉼표·공백 제거."""
    try:
        return float(str(text).replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


# ════════════════════════════════════════════════════════
#  1. rawdata CSV 파싱
# ════════════════════════════════════════════════════════

def _read_sap_xls(path: str) -> list[list[str]]:
    """
    SAP 스프레드시트 내보내기 파일 읽기.

    SAP List→Export→Spreadsheet 가 만드는 .XLS 파일은
    실제로 UTF-16 LE BOM + 탭 구분 텍스트입니다.
    xlrd / openpyxl 로는 읽을 수 없으므로 텍스트로 직접 읽습니다.
    """
    # UTF-16 시도 (SAP 스프레드시트 기본)
    for enc in ("utf-16", "utf-16-le", "utf-16-be"):
        try:
            with open(path, "r", encoding=enc, errors="replace") as f:
                content = f.read()
            rows = []
            for line in content.splitlines():
                cols = [v.strip() for v in line.split("\t")]
                rows.append(cols)
            if any(len(r) > 4 for r in rows):   # 데이터 행이 실제로 있는지 확인
                return rows
        except Exception:
            continue

    # 폴백: openpyxl / xlrd
    try:
        df = pd.read_excel(path, header=None, dtype=str)
        return [_clean_row(row) for _, row in df.iterrows()]
    except Exception:
        pass

    raise ValueError(f"SAP XLS 파일을 읽을 수 없습니다: {path}")


def _is_po_number(v: str) -> bool:
    """SAP 구매문서번호 여부: 8자리 이상 숫자, 4 로 시작 (4500..., 4600... 등)."""
    clean = v.replace(",", "").replace(" ", "")
    return len(clean) >= 8 and clean.isdigit() and clean.startswith("4")


def parse_rawdata(csv_path: str) -> list:
    """
    MR11SHOW 상세 화면 rawdata (SAP XLS 또는 CSV)를 파싱하여 레코드 리스트 반환.

    SAP 스프레드시트 XLS 구조 (탭 구분):
        행 0 : 전표번호  …  col[3] = "5400004827 2026"
        행 1 : 회사코드
        행 2 : 통화
        행 3-4: 빈행
        행 5 : A행 헤더  col[1]=구매문서  col[4]=품목  col[5]=PO일자
                          col[8]=이름1  col[10]=자재  col[11]=Plnt  col[12]=내역  col[13]=OUn
        행 6 : B행 헤더  col[1]=품목(선택번호)  col[2]=구매문서  col[6]=품목
                          col[7]=계정키이름  col[9]=차이수량  col[10]=차이금액
        행 7 : 빈행
        행 8~: 데이터 (A행 / B행 / 빈행 반복)

    A행 식별: col[1] = PO번호(8자리↑, 4로 시작), col[2] = 빈값
    B행 식별: col[2] = PO번호,  col[1] = 순번(숫자)
    동일 PO·품목에 계정키이름이 다른 복수 B행 허용 (관세반제 + 운임반제 등)
    """
    ext = os.path.splitext(csv_path)[1].lower()

    if ext in (".xls", ".xlsx"):
        rows = _read_sap_xls(csv_path)
    else:
        # CSV: 인코딩 자동 감지
        df_raw = None
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                df_raw = pd.read_csv(csv_path, header=None, dtype=str, encoding=enc)
                break
            except (UnicodeDecodeError, Exception):
                continue
        if df_raw is None:
            raise ValueError(f"파일 인코딩을 읽을 수 없습니다: {csv_path}")
        rows = [_clean_row(row) for _, row in df_raw.iterrows()]

    if not rows:
        return []

    # 전표번호 추출 (행 0, col[3], "5400004827 2026" → "5400004827")
    raw_no  = rows[0][3] if len(rows[0]) > 3 else ""
    전표번호 = raw_no.split()[0] if raw_no else ""

    # A행 / B행 분류
    row_a_list: list[tuple[str, list]] = []   # (품목번호, row)
    row_b_list: list[tuple[str, list]] = []   # (품목번호, row)

    for row in rows:
        if len(row) < 8:
            continue
        col1 = row[1] if len(row) > 1 else ""
        col2 = row[2] if len(row) > 2 else ""

        if _is_po_number(col1) and col2 == "":
            # A행: col[1]=PO, col[2]=빈값
            item_no = row[4] if len(row) > 4 else ""
            row_a_list.append((item_no, row))
        elif _is_po_number(col2) and col1.isdigit():
            # B행: col[2]=PO, col[1]=선택순번
            item_no = row[6] if len(row) > 6 else ""
            row_b_list.append((item_no, row))

    # B행을 품목번호 기준 리스트 딕셔너리로 변환
    # (동일 품목에 관세반제·운임반제 등 복수 B행 허용)
    from collections import defaultdict
    b_map: dict[str, list] = defaultdict(list)
    for item_no, row in row_b_list:
        b_map[item_no].append(row)

    # A행 × B행 매핑 → 레코드 생성
    records = []
    for item_no, row_a in row_a_list:
        b_rows = b_map.get(item_no, [])
        if not b_rows:
            continue
        for row_b in b_rows:
            acct = row_b[7] if len(row_b) > 7 else ""
            if not acct:            # 계정키이름 없으면 스킵
                continue
            records.append({
                "전표번호":   전표번호,
                "구매 문서":  row_a[1],
                "품목":       item_no,
                "PO 일자":    row_a[5]  if len(row_a) > 5  else "",
                "계정키이름": acct,
                "이름 1":     row_a[8]  if len(row_a) > 8  else "",
                "차이 수량":  row_b[9]  if len(row_b) > 9  else "",
                "차이 금액":  row_b[10] if len(row_b) > 10 else "",
                "Plnt":       row_a[11] if len(row_a) > 11 else "",
                "내역":       row_a[12] if len(row_a) > 12 else "",
                "OUn":        row_a[13] if len(row_a) > 13 else "",
            })

    return records


# ════════════════════════════════════════════════════════
#  2. 마스터 파일 누적 추가
# ════════════════════════════════════════════════════════

def _get_existing_keys(ws) -> set:
    """
    마스터 워크시트에서 중복 체크용 키 집합을 반환합니다.
    키 = (전표번호, 구매문서, 품목, 계정키이름)
    동일 PO·품목에 운임반제·관세반제 등 복수 계정이 있을 수 있으므로 계정키이름 포함.
    """
    keys = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            keys.add((
                str(row[0]).strip(),
                str(row[1]).strip(),
                str(row[2]).strip() if row[2] else "",
                str(row[4]).strip() if len(row) > 4 and row[4] else "",
            ))
    return keys


def _recalc_total(ws) -> float:
    """차이 금액 컬럼(H열) 합계를 재계산합니다."""
    total = 0.0
    for row in ws.iter_rows(min_row=3, min_col=TOTAL_COL_IDX,
                             max_col=TOTAL_COL_IDX, values_only=True):
        val = row[0]
        if val is not None:
            total += _to_number(str(val))
    return total


def _apply_row_style(ws, row_idx: int, is_new: bool = False):
    """단일 데이터 행에 서식을 적용합니다."""
    fill_color = NEW_ROW_COLOR if is_new else (ALT_ROW_COLOR if row_idx % 2 == 0 else "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)
    thin_border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = Font(size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 16


def _create_master(master_path: str, records: list):
    """마스터 파일이 없을 때 새로 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "반제리스트"

    # 행 1: 합계금액 행 (H열에 합계)
    total = sum(_to_number(r.get("차이 금액", "0")) for r in records)
    for col in range(1, len(MASTER_COLS) + 1):
        ws.cell(row=1, column=col).value = ""
    ws.cell(row=1, column=TOTAL_COL_IDX).value = total
    ws.cell(row=1, column=TOTAL_COL_IDX).number_format = '#,##0'

    # 행 2: 헤더
    header_fill   = PatternFill("solid", fgColor=HEADER_COLOR)
    thin_border   = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )
    for col_idx, col_name in enumerate(MASTER_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=HEADER_FONT, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 데이터 행
    for rec in records:
        row_data = [rec.get(c, "") for c in MASTER_COLS]
        ws.append(row_data)
        _apply_row_style(ws, ws.max_row, is_new=True)

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A3"
    wb.save(master_path)


def append_to_master(records: list, master_path: str) -> dict:
    """
    파싱된 레코드를 마스터 Excel 파일에 누적 추가합니다.

    Returns:
        {"added": int, "skipped": int}
    """
    if not records:
        return {"added": 0, "skipped": 0}

    if not os.path.exists(master_path):
        print(f"  마스터 파일 없음 → 새로 생성: {master_path}")
        _create_master(master_path, records)
        return {"added": len(records), "skipped": 0}

    wb = openpyxl.load_workbook(master_path)
    ws = wb.active

    existing_keys = _get_existing_keys(ws)
    added = 0
    skipped = 0

    for rec in records:
        key = (
            str(rec.get("전표번호", "")).strip(),
            str(rec.get("구매 문서", "")).strip(),
            str(rec.get("품목", "")).strip(),
            str(rec.get("계정키이름", "")).strip(),
        )
        if key in existing_keys:
            skipped += 1
            continue

        row_data = [rec.get(c, "") for c in MASTER_COLS]
        ws.append(row_data)
        _apply_row_style(ws, ws.max_row, is_new=True)
        existing_keys.add(key)
        added += 1

    # 합계금액 재계산
    total = _recalc_total(ws)
    ws.cell(row=1, column=TOTAL_COL_IDX).value = total
    ws.cell(row=1, column=TOTAL_COL_IDX).number_format = '#,##0'

    wb.save(master_path)
    return {"added": added, "skipped": skipped}


# ════════════════════════════════════════════════════════
#  3. 폴더 일괄 처리
# ════════════════════════════════════════════════════════

def process_folder(input_dir: str, master_path: str) -> dict:
    """
    폴더 내 모든 CSV 파일을 처리하여 마스터 파일에 추가합니다.

    Returns:
        {"files": int, "added": int, "skipped": int, "errors": list}
    """
    csv_files = sorted(glob.glob(os.path.join(input_dir, "*.csv")))
    if not csv_files:
        print(f"  [경고] CSV 파일을 찾을 수 없습니다: {input_dir}")
        return {"files": 0, "added": 0, "skipped": 0, "errors": []}

    total_added   = 0
    total_skipped = 0
    errors        = []

    for csv_path in csv_files:
        filename = os.path.basename(csv_path)
        try:
            records = parse_rawdata(csv_path)
            if not records:
                print(f"  [스킵] 파싱된 데이터 없음: {filename}")
                continue

            result = append_to_master(records, master_path)
            total_added   += result["added"]
            total_skipped += result["skipped"]
            print(f"  {filename} → 추가 {result['added']}행, 중복스킵 {result['skipped']}행")

        except Exception as e:
            errors.append({"file": filename, "error": str(e)})
            print(f"  [오류] {filename}: {e}")

    return {
        "files":   len(csv_files),
        "added":   total_added,
        "skipped": total_skipped,
        "errors":  errors,
    }


# ════════════════════════════════════════════════════════
#  4. 메인 실행 흐름
# ════════════════════════════════════════════════════════

def run(input_path: str, master_path: str):
    """MR11SHOW rawdata 처리 전체 파이프라인을 실행합니다."""
    print(f"\n{'='*55}")
    print("  MR11SHOW rawdata 처리기")
    print(f"{'='*55}")
    print(f"  입력  : {input_path}")
    print(f"  마스터: {master_path}")
    print(f"{'='*55}\n")

    if os.path.isdir(input_path):
        print(f"[폴더 모드] CSV 파일 일괄 처리 중...\n")
        result = process_folder(input_path, master_path)
        print(f"\n{'='*55}")
        print(f"  처리 완료")
        print(f"  파일 수  : {result['files']}개")
        print(f"  추가 행수: {result['added']}행")
        print(f"  중복 스킵: {result['skipped']}행")
        if result["errors"]:
            print(f"  오류 발생: {len(result['errors'])}개")
            for e in result["errors"]:
                print(f"    - {e['file']}: {e['error']}")

    elif os.path.isfile(input_path):
        print(f"[파일 모드] {os.path.basename(input_path)} 처리 중...\n")
        records = parse_rawdata(input_path)
        print(f"  파싱 완료: {len(records)}개 품목")

        result = append_to_master(records, master_path)
        print(f"\n{'='*55}")
        print(f"  처리 완료")
        print(f"  추가 행수: {result['added']}행")
        print(f"  중복 스킵: {result['skipped']}행")

    else:
        print(f"[오류] 입력 경로를 찾을 수 없습니다: {input_path}")
        sys.exit(1)

    print(f"{'='*55}")
    print(f"  마스터 파일 저장 완료: {master_path}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="MR11SHOW rawdata CSV → 반제리스트 Excel 누적 추가"
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="rawdata CSV 파일 경로 또는 CSV가 들어있는 폴더 경로"
    )
    parser.add_argument(
        "--master", "-m", required=True,
        help="반제리스트 마스터 Excel 파일 경로 (.xlsx)"
    )
    args = parser.parse_args()

    run(args.input, args.master)
