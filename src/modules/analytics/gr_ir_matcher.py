"""
GR/IR 매칭 엔진
SAP 입고(GR)·송장(IR) 데이터를 PO 번호 기준으로 매칭하여
잔액을 미착품/미확정채무/예외로 자동 분류합니다.
"""

import sys
import os
import argparse
from datetime import datetime, date

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────
HEADER_COLOR  = "1F4E79"   # 헤더 배경색 (진한 파랑)
HEADER_FONT   = "FFFFFF"   # 헤더 글자색 (흰색)
ALT_ROW_COLOR = "D6E4F0"   # 짝수 행 배경색 (연한 파랑)

# 분류 레이블
CLS_MATCHED    = "완전매칭"
CLS_GOODS_RECV = "미착품"       # GR 있음, IR 없음
CLS_ACCRUED    = "미확정채무"   # IR 있음, GR 없음
CLS_DIFF_AMT   = "예외_금액차이"
CLS_DIFF_QTY   = "예외_수량차이"

# Aging 구간 레이블
AGING_LABELS = {
    "30일이하": (0, 30),
    "31~60일":  (31, 60),
    "61~90일":  (61, 90),
    "90일초과": (91, None),
}
# ──────────────────────────────────────────────────────

# ── 필수 컬럼 매핑 (SAP 한국어 기본 컬럼명) ────────────
# 실제 SAP 추출 파일의 컬럼명이 다를 경우 이 딕셔너리를 수정하세요.
GR_COL_MAP = {
    "po_no":    "구매오더",   # PO 번호
    "po_date":  "문서날짜",   # PO 생성일 (Aging 기준)
    "mat_code": "자재",       # 자재코드
    "qty":      "입고수량",   # 입고 수량
    "amount":   "금액",       # 입고 금액
    "currency": "통화",       # 통화 코드
}

IR_COL_MAP = {
    "po_no":    "구매오더",   # PO 번호
    "po_date":  "전기일",     # 송장 전기일
    "mat_code": "자재",       # 자재코드
    "qty":      "수량",       # 송장 수량
    "amount":   "금액",       # 송장 금액
    "currency": "통화",       # 통화 코드
}
# ──────────────────────────────────────────────────────


# ════════════════════════════════════════════════════════
#  1. 데이터 로드
# ════════════════════════════════════════════════════════

def _read_excel(path: str) -> pd.DataFrame:
    """Excel 파일을 읽어 DataFrame으로 반환합니다. SAP 상단 메타헤더를 자동 탐지합니다."""
    df_raw = pd.read_excel(path, header=None, dtype=str)

    # SAP 메타헤더 건너뛰기: 비어있지 않은 셀이 3개 이상인 첫 행을 헤더로 사용
    start_row = 0
    for i, row in df_raw.iterrows():
        non_empty = row.dropna().astype(str).str.strip().replace("", pd.NA).dropna()
        if len(non_empty) >= 3:
            start_row = i
            break

    df = df_raw.iloc[start_row:].reset_index(drop=True)
    df.columns = df.iloc[0].fillna("").astype(str).str.strip()
    df = df.iloc[1:].reset_index(drop=True)

    # 완전히 빈 행/열 제거
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.dropna(axis=1, how="all")
    df = df[~df.apply(lambda r: r.astype(str).str.strip().eq("").all(), axis=1)]
    return df.reset_index(drop=True)


def _to_numeric(series: pd.Series) -> pd.Series:
    """문자열 숫자 컬럼을 float으로 변환합니다. 쉼표·공백을 제거합니다."""
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    ).fillna(0.0)


def load_gr_data(path: str) -> pd.DataFrame:
    """
    GR(입고) Excel을 로드하고 표준 컬럼명으로 정규화합니다.

    Returns:
        컬럼: po_no, po_date, mat_code, gr_qty, gr_amount, currency
    """
    df = _read_excel(path)
    col = GR_COL_MAP

    missing = [v for v in col.values() if v not in df.columns]
    if missing:
        available = list(df.columns)
        raise ValueError(
            f"[GR 파일] 필수 컬럼을 찾을 수 없습니다: {missing}\n"
            f"파일의 실제 컬럼: {available}\n"
            f"gr_ir_matcher.py 상단의 GR_COL_MAP을 수정하세요."
        )

    result = pd.DataFrame({
        "po_no":     df[col["po_no"]].astype(str).str.strip(),
        "po_date":   pd.to_datetime(df[col["po_date"]], errors="coerce"),
        "mat_code":  df[col["mat_code"]].astype(str).str.strip(),
        "gr_qty":    _to_numeric(df[col["qty"]]),
        "gr_amount": _to_numeric(df[col["amount"]]),
        "currency":  df[col["currency"]].astype(str).str.strip(),
    })
    print(f"  GR 데이터 로드 완료: {len(result)}행 ({path})")
    return result


def load_ir_data(path: str) -> pd.DataFrame:
    """
    IR(송장) Excel을 로드하고 표준 컬럼명으로 정규화합니다.

    Returns:
        컬럼: po_no, po_date, mat_code, ir_qty, ir_amount, currency
    """
    df = _read_excel(path)
    col = IR_COL_MAP

    missing = [v for v in col.values() if v not in df.columns]
    if missing:
        available = list(df.columns)
        raise ValueError(
            f"[IR 파일] 필수 컬럼을 찾을 수 없습니다: {missing}\n"
            f"파일의 실제 컬럼: {available}\n"
            f"gr_ir_matcher.py 상단의 IR_COL_MAP을 수정하세요."
        )

    result = pd.DataFrame({
        "po_no":     df[col["po_no"]].astype(str).str.strip(),
        "po_date":   pd.to_datetime(df[col["po_date"]], errors="coerce"),
        "mat_code":  df[col["mat_code"]].astype(str).str.strip(),
        "ir_qty":    _to_numeric(df[col["qty"]]),
        "ir_amount": _to_numeric(df[col["amount"]]),
        "currency":  df[col["currency"]].astype(str).str.strip(),
    })
    print(f"  IR 데이터 로드 완료: {len(result)}행 ({path})")
    return result


# ════════════════════════════════════════════════════════
#  2. 매칭 엔진
# ════════════════════════════════════════════════════════

def match_gr_ir(gr_df: pd.DataFrame, ir_df: pd.DataFrame,
                tolerance: float = 1.0) -> pd.DataFrame:
    """
    PO 번호 기준으로 GR·IR을 매칭하고 잔액을 계산하여 분류합니다.

    Args:
        gr_df:     load_gr_data() 결과
        ir_df:     load_ir_data() 결과
        tolerance: 완전매칭 허용 오차 금액 (기본 1원)

    Returns:
        분류 결과 DataFrame. 컬럼:
            po_no, po_date, gr_qty, gr_amount, ir_qty, ir_amount,
            잔액, 분류, currency
    """
    # PO별 GR 합계
    gr_agg = gr_df.groupby("po_no", as_index=False).agg(
        po_date  =("po_date",   "min"),
        gr_qty   =("gr_qty",    "sum"),
        gr_amount=("gr_amount", "sum"),
        currency =("currency",  "first"),
    )

    # PO별 IR 합계
    ir_agg = ir_df.groupby("po_no", as_index=False).agg(
        ir_qty   =("ir_qty",    "sum"),
        ir_amount=("ir_amount", "sum"),
    )

    # outer join — 어느 한쪽에만 있는 PO도 포함
    merged = pd.merge(gr_agg, ir_agg, on="po_no", how="outer")
    merged["gr_qty"]    = merged["gr_qty"].fillna(0.0)
    merged["gr_amount"] = merged["gr_amount"].fillna(0.0)
    merged["ir_qty"]    = merged["ir_qty"].fillna(0.0)
    merged["ir_amount"] = merged["ir_amount"].fillna(0.0)
    merged["currency"]  = merged["currency"].fillna("")

    # 잔액 계산
    merged["잔액"] = merged["gr_amount"] - merged["ir_amount"]

    # 분류
    merged["분류"] = merged.apply(
        lambda r: _classify(r, tolerance), axis=1
    )

    # 컬럼 순서 정리
    result = merged[[
        "po_no", "po_date",
        "gr_qty", "gr_amount",
        "ir_qty", "ir_amount",
        "잔액", "분류", "currency",
    ]].rename(columns={
        "po_no":     "PO번호",
        "po_date":   "PO일자",
        "gr_qty":    "GR수량",
        "gr_amount": "GR금액",
        "ir_qty":    "IR수량",
        "ir_amount": "IR금액",
    })

    return result.reset_index(drop=True)


def _classify(row: pd.Series, tolerance: float) -> str:
    """단일 PO 행의 분류 레이블을 반환합니다."""
    gr_amt = row["gr_amount"]
    ir_amt = row["ir_amount"]
    gr_qty = row["gr_qty"]
    ir_qty = row["ir_qty"]
    balance = gr_amt - ir_amt

    if abs(balance) <= tolerance:
        return CLS_MATCHED

    if gr_amt > 0 and ir_amt == 0:
        return CLS_GOODS_RECV      # 미착품

    if ir_amt > 0 and gr_amt == 0:
        return CLS_ACCRUED         # 미확정채무

    # 금액 차이와 수량 차이가 모두 있으면 금액 차이 우선
    if abs(gr_qty - ir_qty) > 0:
        return CLS_DIFF_QTY

    return CLS_DIFF_AMT


# ════════════════════════════════════════════════════════
#  3. Aging 계산
# ════════════════════════════════════════════════════════

def add_aging(df: pd.DataFrame, base_date: date = None) -> pd.DataFrame:
    """
    PO일자 기준으로 경과일수와 Aging 구간을 추가합니다.

    Args:
        df:        match_gr_ir() 결과
        base_date: 기준일 (기본값: 오늘)

    Returns:
        경과일수, Aging구간 컬럼이 추가된 DataFrame
    """
    if base_date is None:
        base_date = date.today()

    base_ts = pd.Timestamp(base_date)

    df = df.copy()
    df["경과일수"] = (base_ts - df["PO일자"]).dt.days.fillna(-1).astype(int)

    def _aging_label(days: int) -> str:
        if days < 0:
            return "날짜없음"
        for label, (lo, hi) in AGING_LABELS.items():
            if hi is None and days >= lo:
                return label
            if hi is not None and lo <= days <= hi:
                return label
        return "날짜없음"

    df["Aging구간"] = df["경과일수"].apply(_aging_label)
    return df


# ════════════════════════════════════════════════════════
#  4. 리포트 생성
# ════════════════════════════════════════════════════════

def export_report(df: pd.DataFrame, output_path: str) -> str:
    """
    매칭 결과를 5개 시트로 분리하여 Excel 리포트를 저장합니다.

    시트 구성:
        완전매칭 / 미착품 / 미확정채무 / 예외건 / 전체

    Args:
        df:          add_aging() 이후의 DataFrame
        output_path: 저장 경로

    Returns:
        저장된 파일 경로
    """
    exception_labels = {CLS_DIFF_AMT, CLS_DIFF_QTY}

    sheets = {
        "완전매칭":   df[df["분류"] == CLS_MATCHED].copy(),
        "미착품":     df[df["분류"] == CLS_GOODS_RECV].copy(),
        "미확정채무": df[df["분류"] == CLS_ACCRUED].copy(),
        "예외건":     df[df["분류"].isin(exception_labels)].copy(),
        "전체":       df.copy(),
    }

    # 예외건 시트에 검토의견 컬럼 추가
    sheets["예외건"]["검토의견"] = ""

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            if sheet_df.empty:
                # 빈 시트도 헤더만 넣어서 생성
                sheet_df = pd.DataFrame(columns=df.columns)
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 서식 적용
    wb = openpyxl.load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _apply_styles(ws)
        _add_summary_row(ws, sheet_name, sheets)
    wb.save(output_path)

    return output_path


def _apply_styles(ws) -> None:
    """워크시트에 표준 서식을 적용합니다."""
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    thin_border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, color=HEADER_FONT, size=10)
            elif row_idx % 2 == 0:
                cell.fill = alt_fill
                cell.font = Font(size=10)
            else:
                cell.font = Font(size=10)

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len   = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # 행 높이
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 16

    # 틀 고정
    ws.freeze_panes = "A2"


def _add_summary_row(ws, sheet_name: str, sheets: dict) -> None:
    """전체 시트 맨 아래에 건수 요약 행을 추가합니다."""
    if sheet_name != "전체":
        return

    summary_fill = PatternFill("solid", fgColor="FFF2CC")  # 연한 노랑
    summary_font = Font(bold=True, size=10)

    labels = [
        ("완전매칭",   CLS_MATCHED),
        ("미착품",     CLS_GOODS_RECV),
        ("미확정채무", CLS_ACCRUED),
        ("예외건",     None),
    ]

    ws.append([""] * ws.max_column)  # 빈 구분 행
    for label, cls in labels:
        if cls is not None:
            cnt = len(sheets[label])
            amt = sheets[label]["GR금액"].sum() - sheets[label]["IR금액"].sum()
        else:
            cnt = len(sheets["예외건"])
            amt = sheets["예외건"]["잔액"].sum()

        summary = [f"[요약] {label}", f"{cnt}건", f"잔액 합계: {amt:,.0f}원"] + \
                  [""] * (ws.max_column - 3)
        ws.append(summary)
        for cell in ws[ws.max_row]:
            cell.fill = summary_fill
            cell.font = summary_font
            cell.alignment = Alignment(horizontal="left", vertical="center")


# ════════════════════════════════════════════════════════
#  5. 메인 실행 흐름
# ════════════════════════════════════════════════════════

def run(gr_path: str, ir_path: str,
        output_path: str = None, tolerance: float = 1.0) -> str:
    """
    GR/IR 매칭 전체 파이프라인을 실행합니다.

    Returns:
        저장된 리포트 파일 경로
    """
    if output_path is None:
        base_dir = os.path.dirname(os.path.abspath(gr_path))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(base_dir, f"GRIR_매칭결과_{timestamp}.xlsx")

    print("\n[1/4] 데이터 로드 중...")
    gr_df = load_gr_data(gr_path)
    ir_df = load_ir_data(ir_path)

    print("[2/4] GR/IR 매칭 중...")
    matched_df = match_gr_ir(gr_df, ir_df, tolerance=tolerance)

    print("[3/4] Aging 계산 중...")
    result_df = add_aging(matched_df)

    print(f"[4/4] 리포트 저장 중: {output_path}")
    export_report(result_df, output_path)

    # 결과 요약 출력
    print("\n" + "=" * 50)
    print("매칭 결과 요약")
    print("=" * 50)
    counts = result_df["분류"].value_counts()
    total  = len(result_df)
    for label in [CLS_MATCHED, CLS_GOODS_RECV, CLS_ACCRUED, CLS_DIFF_AMT, CLS_DIFF_QTY]:
        cnt = counts.get(label, 0)
        print(f"  {label:<12}: {cnt:>5}건")
    print(f"  {'합계':<12}: {total:>5}건")
    print("=" * 50)

    goods_recv_total = result_df[result_df["분류"] == CLS_GOODS_RECV]["잔액"].sum()
    accrued_total    = result_df[result_df["분류"] == CLS_ACCRUED]["잔액"].sum()
    print(f"  미착품 잔액 합계   : {goods_recv_total:>15,.0f}원")
    print(f"  미확정채무 잔액 합계: {accrued_total:>14,.0f}원")
    print("=" * 50)
    print(f"\n저장 위치: {output_path}\n")

    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="SAP GR/IR 매칭 자동화 — 재고매입결산용"
    )
    parser.add_argument("--gr",        required=True,  help="GR(입고) Excel 파일 경로")
    parser.add_argument("--ir",        required=True,  help="IR(송장) Excel 파일 경로")
    parser.add_argument("--output",    default=None,   help="결과 저장 경로 (생략 시 자동 생성)")
    parser.add_argument("--tolerance", type=float, default=1.0,
                        help="완전매칭 허용 오차 금액 (기본: 1원)")
    args = parser.parse_args()

    try:
        run(args.gr, args.ir, args.output, args.tolerance)
    except ValueError as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
