"""
SAP Excel 자동 정리 스크립트
SAP에서 다운로드한 회계/비용 데이터 Excel을 깔끔하게 정리합니다.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime


# ── 설정 ──────────────────────────────────────────────
HEADER_COLOR = "1F4E79"   # 헤더 배경색 (진한 파랑)
HEADER_FONT  = "FFFFFF"   # 헤더 글자색 (흰색)
ALT_ROW_COLOR = "D6E4F0"  # 짝수 행 배경색 (연한 파랑)
# ──────────────────────────────────────────────────────


def find_data_start(df_raw: pd.DataFrame) -> int:
    """SAP 헤더 메타데이터를 건너뛰고 실제 데이터가 시작되는 행 번호를 찾습니다."""
    for i, row in df_raw.iterrows():
        # 비어있지 않은 셀이 3개 이상이면 데이터 행으로 판단
        non_empty = row.dropna().astype(str).str.strip().replace("", pd.NA).dropna()
        if len(non_empty) >= 3:
            return i
    return 0


def clean_sap_excel(input_path: str, output_path: str = None) -> str:
    """
    SAP Excel 파일을 읽어 정리 후 새 파일로 저장합니다.

    Args:
        input_path: SAP에서 받은 원본 Excel 경로
        output_path: 저장할 경로 (없으면 자동 생성)

    Returns:
        저장된 파일 경로
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"{base}_정리완료_{timestamp}{ext}"

    print(f"[1/4] 파일 읽는 중: {input_path}")
    df_raw = pd.read_excel(input_path, header=None, dtype=str)

    # ── 1. 실제 데이터 시작 행 찾기 ──────────────────
    print("[2/4] SAP 헤더 제거 중...")
    start_row = find_data_start(df_raw)
    df = df_raw.iloc[start_row:].reset_index(drop=True)

    # 첫 번째 행을 컬럼명으로 설정
    df.columns = df.iloc[0].fillna("").astype(str).str.strip()
    df = df.iloc[1:].reset_index(drop=True)

    # ── 2. 불필요한 행/열 제거 ───────────────────────
    print("[3/4] 빈 행/열 및 불필요한 데이터 제거 중...")

    # 빈 열 제거
    df = df.loc[:, df.columns.str.strip() != ""]
    df = df.dropna(axis=1, how="all")

    # 빈 행 제거 (모든 셀이 비거나 NaN인 행)
    df = df.dropna(how="all")
    df = df[~df.apply(lambda r: r.astype(str).str.strip().eq("").all(), axis=1)]

    # SAP 합계/소계 행 제거 (선택: 필요 시 주석 처리)
    sap_footer_keywords = ["합계", "소계", "Grand Total", "Total", "Subtotal"]
    mask = df.apply(
        lambda r: r.astype(str).str.contains("|".join(sap_footer_keywords), na=False).any(),
        axis=1
    )
    df_no_totals = df[~mask]
    # 합계 행이 너무 많이 제거되면 원본 유지
    if len(df_no_totals) > len(df) * 0.5:
        df = df_no_totals

    df = df.reset_index(drop=True)

    # ── 3. Excel 저장 및 서식 적용 ───────────────────
    print(f"[4/4] 정리된 파일 저장 중: {output_path}")
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    _apply_styles(ws)

    wb.save(output_path)

    rows, cols = df.shape
    print(f"\n완료! {rows}행 x {cols}열 데이터가 저장되었습니다.")
    print(f"저장 위치: {output_path}")
    return output_path


def _apply_styles(ws):
    """워크시트에 서식을 적용합니다."""
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    thin_border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

            if row_idx == 1:
                # 헤더 행
                cell.fill = header_fill
                cell.font = Font(bold=True, color=HEADER_FONT, size=10)
            elif row_idx % 2 == 0:
                # 짝수 행 (연한 배경)
                cell.fill = alt_fill
                cell.font = Font(size=10)
            else:
                cell.font = Font(size=10)

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # 행 높이
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 16

    # 틀 고정 (헤더 행 고정)
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python sap_excel_cleaner.py <SAP_Excel_파일경로> [저장경로]")
        print("예시:   python sap_excel_cleaner.py C:/Downloads/SAP_report.xlsx")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    clean_sap_excel(input_file, output_file)
